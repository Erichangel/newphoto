import openpyxl
wb = openpyxl.load_workbook('k:/时光印记1.0/测试流程_补充版.xlsx')
ws = wb.active
for i in range(1, ws.max_row+1):
    r = []
    for j in range(1, ws.max_column+1):
        v = ws.cell(row=i, column=j).value
        if v is not None:
            r.append(str(v)[:80])
    if r:
        print(f'Row{i}: {r}')
